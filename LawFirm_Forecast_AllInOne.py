"""
Law Firm Revenue Forecasting — All-in-One Script
================================================
Input  : LawFirm_Synthetic_3Y.xlsx
Outputs: LawFirm_Forecast_3Y.png
         LawFirm_Forecast_Results.xlsx  (3 sheets: Summary, Monthly Detail, Win Count)
"""

import warnings
warnings.filterwarnings('ignore')

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
# 1. LOAD DATA
# ══════════════════════════════════════════════════════════════════════════════

df = pd.read_excel('LawFirm_Synthetic_3Y.xlsx', sheet_name='Monthly Revenue', header=3)
df = df.rename(columns={'Date': 'ds', 'Revenue (EUR)': 'y'})
df = df[['ds', 'y']].dropna().sort_values('ds').reset_index(drop=True)
df['ds'] = pd.to_datetime(df['ds'])

TRAIN_END = '2023-12-01'
train_df = df[df['ds'] <= TRAIN_END].reset_index(drop=True)
test_df  = df[df['ds'] >  TRAIN_END].reset_index(drop=True)
N_TRAIN  = len(train_df)
N_TEST   = len(test_df)

print(f"Dataset : {len(df)} months ({df['ds'].min().strftime('%b %Y')} -> {df['ds'].max().strftime('%b %Y')})")
print(f"Train   : {N_TRAIN} months | Test: {N_TEST} months\n")

# ══════════════════════════════════════════════════════════════════════════════
# 2. PROPHET
# ══════════════════════════════════════════════════════════════════════════════

from prophet import Prophet

print("[1/4] Training Prophet...")

prophet_model = Prophet(
    yearly_seasonality=True,
    weekly_seasonality=False,
    daily_seasonality=False,
    seasonality_mode='multiplicative'
)
prophet_model.fit(train_df)

future   = prophet_model.make_future_dataframe(periods=N_TEST, freq='MS')
forecast = prophet_model.predict(future)

prophet_test_pred = forecast.tail(N_TEST)['yhat'].values
prophet_fitted    = forecast.head(N_TRAIN)['yhat'].values

prophet_mae  = np.mean(np.abs(test_df['y'].values - prophet_test_pred))
prophet_mape = np.mean(np.abs((test_df['y'].values - prophet_test_pred) / test_df['y'].values)) * 100
prophet_rmse = np.sqrt(np.mean((test_df['y'].values - prophet_test_pred) ** 2))

print(f"   MAE: {prophet_mae:,.0f}  MAPE: {prophet_mape:.1f}%  RMSE: {prophet_rmse:,.0f}")

# ══════════════════════════════════════════════════════════════════════════════
# 3. LIGHTGBM VIA DARTS
# ══════════════════════════════════════════════════════════════════════════════

from darts import TimeSeries
from darts.models import LightGBMModel
from darts.metrics import mae as d_mae, mape as d_mape, rmse as d_rmse

print("[2/4] Training LightGBM...")

LAGS = [-1, -2, -3, -6, -12]

series_train = TimeSeries.from_dataframe(train_df, time_col='ds', value_cols='y', freq='MS')
series_test  = TimeSeries.from_dataframe(test_df,  time_col='ds', value_cols='y', freq='MS')

lgbm_model = LightGBMModel(
    lags=LAGS,
    output_chunk_length=6,
    n_estimators=200,
    num_leaves=15,
    min_child_samples=3,
    learning_rate=0.05,
    verbose=-1,
    random_state=42
)
lgbm_model.fit(series_train)
lgbm_pred = lgbm_model.predict(N_TEST)

lgbm_mae  = float(d_mae(series_test,  lgbm_pred))
lgbm_mape = float(d_mape(series_test, lgbm_pred))
lgbm_rmse = float(d_rmse(series_test, lgbm_pred))

print(f"   MAE: {lgbm_mae:,.0f}  MAPE: {lgbm_mape:.1f}%  RMSE: {lgbm_rmse:,.0f}")

lgbm_fitted_df = lgbm_model.historical_forecasts(
    series_train, forecast_horizon=1, retrain=False, verbose=False
).to_dataframe().reset_index()
lgbm_fitted_df.columns = ['ds', 'yhat']

lgbm_pred_df = lgbm_pred.to_dataframe().reset_index()
lgbm_pred_df.columns = ['ds', 'yhat']

winner = "Prophet" if prophet_mae < lgbm_mae else "LightGBM"

# Terminal summary
print(f"\n{'='*62}")
print(f"  MODEL COMPARISON - out-of-sample test ({N_TEST} months)")
print(f"{'='*62}")
print(f"  {'Model':<14} {'MAE':>10} {'MAPE':>8} {'RMSE':>10}")
print(f"  {'-'*56}")
print(f"  {'Prophet':<14} {prophet_mae:>10,.0f} {prophet_mape:>7.1f}% {prophet_rmse:>10,.0f}  {'<- best' if winner=='Prophet' else ''}")
print(f"  {'LightGBM':<14} {lgbm_mae:>10,.0f} {lgbm_mape:>7.1f}% {lgbm_rmse:>10,.0f}  {'<- best' if winner=='LightGBM' else ''}")
print(f"{'='*62}")
print(f"\n  {'Month':<12} {'Actual':>10} {'Prophet':>10} {'LightGBM':>10} {'Err P':>7} {'Err L':>7}")
print(f"  {'-'*60}")
for i in range(N_TEST):
    m   = test_df['ds'].iloc[i].strftime('%b %Y')
    act = test_df['y'].iloc[i]
    p   = prophet_test_pred[i]
    l   = lgbm_pred_df['yhat'].iloc[i]
    print(f"  {m:<12} {act:>10,.0f} {p:>10,.0f} {l:>10,.0f} {abs(act-p)/act*100:>6.0f}% {abs(act-l)/act*100:>6.0f}%")

# ══════════════════════════════════════════════════════════════════════════════
# 4. CHART
# ══════════════════════════════════════════════════════════════════════════════

print("\n[3/4] Generating chart...")

C       = dict(actual='#1F4E79', prophet='#E74C3C', lgbm='#27AE60', vline='#7F8C8D')
fmt_eur = plt.FuncFormatter(lambda x, _: f'{x:,.0f}')

fig = plt.figure(figsize=(15, 11))
gs  = gridspec.GridSpec(2, 3, figure=fig, hspace=0.42, wspace=0.35)
ax1 = fig.add_subplot(gs[0, :])
ax2 = fig.add_subplot(gs[1, 0:2])
ax3 = fig.add_subplot(gs[1, 2])

# ax1 — full series
ax1.axvspan(test_df['ds'].min(), test_df['ds'].max(), alpha=0.07, color='orange', label='Test period')
ax1.plot(df['ds'], df['y'], color=C['actual'], lw=2.5, marker='o', ms=4, zorder=5, label='Actual')
ax1.plot(train_df['ds'], prophet_fitted, color=C['prophet'], lw=1.5, ls=':', alpha=0.5)
ax1.plot(test_df['ds'], prophet_test_pred, color=C['prophet'], lw=2, ls='--', label='Prophet (test)')
ax1.fill_between(forecast['ds'], forecast['yhat_lower'], forecast['yhat_upper'], alpha=0.10, color=C['prophet'])
ax1.plot(lgbm_fitted_df['ds'], lgbm_fitted_df['yhat'], color=C['lgbm'], lw=1.5, ls=':', alpha=0.5)
ax1.plot(lgbm_pred_df['ds'], lgbm_pred_df['yhat'], color=C['lgbm'], lw=2, ls='--', marker='s', ms=4, label='LightGBM (test)')
ax1.axvline(x=test_df['ds'].min(), color=C['vline'], ls=':', lw=1.5)
ax1.set_title('Law Firm Revenue — Prophet vs LightGBM | Train: 2022–2023 | Test: 2024–mid 2025', fontsize=12, fontweight='bold')
ax1.set_xlabel('Month')
ax1.set_ylabel('Revenue (EUR)')
ax1.legend(fontsize=9)
ax1.grid(True, alpha=0.3)
ax1.yaxis.set_major_formatter(fmt_eur)

# ax2 — test zoom
ax2.plot(test_df['ds'], test_df['y'], color=C['actual'], lw=2.5, marker='o', ms=7, label='Actual', zorder=5)
ax2.plot(test_df['ds'], prophet_test_pred, color=C['prophet'], lw=2, ls='--', marker='^', ms=7,
         label=f'Prophet  MAE {prophet_mae:,.0f} | MAPE {prophet_mape:.1f}%')
ax2.plot(lgbm_pred_df['ds'], lgbm_pred_df['yhat'], color=C['lgbm'], lw=2, ls='--', marker='s', ms=7,
         label=f'LightGBM MAE {lgbm_mae:,.0f} | MAPE {lgbm_mape:.1f}%')
ax2.set_title('Out-of-Sample Test Period (2024 – mid 2025)', fontsize=11, fontweight='bold')
ax2.set_xlabel('Month')
ax2.set_ylabel('Revenue (EUR)')
ax2.legend(fontsize=8)
ax2.grid(True, alpha=0.3)
ax2.yaxis.set_major_formatter(fmt_eur)
ax2.tick_params(axis='x', rotation=30)

# ax3 — metric bars
p_vals = [prophet_mae, prophet_rmse]
l_vals = [lgbm_mae,    lgbm_rmse]
x = np.arange(2)
w = 0.32
bars_p = ax3.bar(x - w/2, p_vals, w, color=C['prophet'], label='Prophet',  edgecolor='white')
bars_l = ax3.bar(x + w/2, l_vals, w, color=C['lgbm'],    label='LightGBM', edgecolor='white')
for bar in list(bars_p) + list(bars_l):
    ax3.text(bar.get_x() + bar.get_width()/2, bar.get_height() + max(max(p_vals), max(l_vals)) * 0.01,
             f'{bar.get_height():,.0f}', ha='center', va='bottom', fontsize=8, fontweight='bold')
ax3.set_title('Error Metrics (lower = better)', fontsize=11, fontweight='bold')
ax3.set_xticks(x)
ax3.set_xticklabels(['MAE', 'RMSE'])
ax3.set_ylabel('EUR')
ax3.legend(fontsize=9)
ax3.grid(True, alpha=0.3, axis='y')
ax3.yaxis.set_major_formatter(fmt_eur)

fig.text(0.5, 0.01, f'Winner (MAE): {winner}  |  Train: {N_TRAIN}m  |  Test: {N_TEST}m  |  LightGBM lags: {LAGS}',
         ha='center', fontsize=10, fontweight='bold', color='#2C3E50',
         bbox=dict(boxstyle='round,pad=0.4', facecolor='#ECF0F1', edgecolor='#BDC3C7'))

plt.savefig('LawFirm_Forecast_3Y.png', dpi=150, bbox_inches='tight')
print("Saved: LawFirm_Forecast_3Y.png")

# ══════════════════════════════════════════════════════════════════════════════
# 5. EXCEL RESULTS REPORT
# ══════════════════════════════════════════════════════════════════════════════

print("[4/4] Building Excel report...")

# Helpers
def make_border():
    s = Side(style='thin', color='BDC3C7')
    return Border(left=s, right=s, top=s, bottom=s)

DARK_BLUE  = '1F4E79'
MID_BLUE   = '2E75B6'
LIGHT_BLUE = 'D6E4F0'
ALT_BLUE   = 'EBF3FB'
GREEN      = '27AE60'
GREEN_BG   = 'E8F5E9'
RED        = 'C0392B'
GRAY       = '7F8C8D'
WHITE      = 'FFFFFF'

def hdr(ws, row, col, val, bg=DARK_BLUE, size=11):
    c = ws.cell(row, col, val)
    c.font      = Font(name='Arial', bold=True, color=WHITE, size=size)
    c.fill      = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border    = make_border()
    return c

def dat(ws, row, col, val, fmt=None, bold=False, bg=None, color='000000'):
    c = ws.cell(row, col, val)
    c.font      = Font(name='Arial', bold=bold, size=10, color=color)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border    = make_border()
    if fmt: c.number_format = fmt
    if bg:  c.fill = PatternFill('solid', start_color=bg)
    return c

EUR = '#,##0'

wb2 = Workbook()

# ── Sheet 1: Summary ──────────────────────────────────────────────────────────
ws1 = wb2.active
ws1.title = 'Summary'
ws1.sheet_view.showGridLines = False

ws1.merge_cells('B2:H2')
ws1['B2'] = 'Law Firm Revenue Forecasting — Model Comparison'
ws1['B2'].font      = Font(name='Arial', bold=True, size=16, color=DARK_BLUE)
ws1['B2'].alignment = Alignment(horizontal='left', vertical='center')
ws1.row_dimensions[2].height = 28

ws1.merge_cells('B3:H3')
ws1['B3'] = f'Out-of-sample validation  |  Train: Jan 2022 – Dec 2023 ({N_TRAIN}m)  |  Test: Jan 2024 – Jun 2025 ({N_TEST}m)'
ws1['B3'].font      = Font(name='Arial', size=10, color=GRAY, italic=True)
ws1['B3'].alignment = Alignment(horizontal='left', vertical='center')
ws1.row_dimensions[3].height = 18
ws1.row_dimensions[4].height = 10

for col, h in enumerate(['Model', 'MAE (EUR)', 'MAPE', 'RMSE (EUR)', 'Seasonality', 'Lags Used', 'Result'], 2):
    hdr(ws1, 5, col, h)
ws1.row_dimensions[5].height = 22

p_bg = GREEN_BG if winner == 'Prophet'  else ALT_BLUE
l_bg = GREEN_BG if winner == 'LightGBM' else ALT_BLUE

dat(ws1, 6, 2, 'Prophet',              bold=True, bg=p_bg)
dat(ws1, 6, 3, prophet_mae,  fmt=EUR,  bold=True, bg=p_bg, color=GREEN if winner=='Prophet' else '000000')
dat(ws1, 6, 4, prophet_mape/100, fmt='0.0%', bg=p_bg)
dat(ws1, 6, 5, prophet_rmse, fmt=EUR,  bg=p_bg)
dat(ws1, 6, 6, 'Multiplicative yearly', bg=p_bg)
dat(ws1, 6, 7, 'Built-in',             bg=p_bg)
dat(ws1, 6, 8, 'WINNER' if winner=='Prophet' else '', bold=True, bg=p_bg,
    color=GREEN if winner=='Prophet' else '000000')
ws1.row_dimensions[6].height = 22

dat(ws1, 7, 2, 'LightGBM',             bold=True, bg=l_bg)
dat(ws1, 7, 3, lgbm_mae,    fmt=EUR,   bold=True, bg=l_bg, color=GREEN if winner=='LightGBM' else '000000')
dat(ws1, 7, 4, lgbm_mape/100,  fmt='0.0%', bg=l_bg)
dat(ws1, 7, 5, lgbm_rmse,   fmt=EUR,   bg=l_bg)
dat(ws1, 7, 6, 'Lag-based (ML)',        bg=l_bg)
dat(ws1, 7, 7, str(LAGS),              bg=l_bg)
dat(ws1, 7, 8, 'WINNER' if winner=='LightGBM' else '', bold=True, bg=l_bg,
    color=GREEN if winner=='LightGBM' else '000000')
ws1.row_dimensions[7].height = 22
ws1.row_dimensions[8].height = 12

ws1.merge_cells('B9:H9')
ws1['B9'] = 'Key Finding'
ws1['B9'].font      = Font(name='Arial', bold=True, size=11, color=DARK_BLUE)
ws1['B9'].fill      = PatternFill('solid', start_color=LIGHT_BLUE)
ws1['B9'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws1['B9'].border    = make_border()
ws1.row_dimensions[9].height = 20

finding = (
    f'{winner} wins with MAE €{min(prophet_mae, lgbm_mae):,.0f} vs €{max(prophet_mae, lgbm_mae):,.0f}. '
    f'Prophet captures the strong yearly seasonal pattern (multiplicative mode) built into this dataset. '
    f'LightGBM with lag-12 needs more training data to fully exploit annual seasonality — '
    f'with 3+ years of history its performance is expected to improve significantly.'
)
ws1.merge_cells('B10:H12')
ws1['B10'] = finding
ws1['B10'].font      = Font(name='Arial', size=10)
ws1['B10'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
ws1['B10'].border    = make_border()
for r in [10, 11, 12]:
    ws1.row_dimensions[r].height = 18

for col, w in {'B': 14, 'C': 13, 'D': 10, 'E': 13, 'F': 22, 'G': 22, 'H': 12}.items():
    ws1.column_dimensions[col].width = w

# ── Sheet 2: Monthly Detail ───────────────────────────────────────────────────
ws2 = wb2.create_sheet('Monthly Detail')
ws2.sheet_view.showGridLines = False

ws2.merge_cells('B2:J2')
ws2['B2'] = 'Monthly Predictions vs Actuals — Test Period'
ws2['B2'].font      = Font(name='Arial', bold=True, size=14, color=DARK_BLUE)
ws2['B2'].alignment = Alignment(horizontal='left', vertical='center')
ws2.row_dimensions[2].height = 26
ws2.row_dimensions[3].height = 10

for col, h in enumerate(['Month', 'Actual', 'Prophet', 'LightGBM',
                          'Err Prophet', 'Err LightGBM', 'Better Model',
                          'Abs Diff', 'Diff %'], 2):
    hdr(ws2, 4, col, h)
ws2.row_dimensions[4].height = 22

p_wins = 0
l_wins = 0
for i in range(N_TEST):
    r    = i + 5
    m    = test_df['ds'].iloc[i].strftime('%b %Y')
    act  = float(test_df['y'].iloc[i])
    p    = float(prophet_test_pred[i])
    l    = float(lgbm_pred_df['yhat'].iloc[i])
    ep   = abs(act - p) / act
    el   = abs(act - l) / act
    better = 'Prophet' if ep <= el else 'LightGBM'
    if better == 'Prophet': p_wins += 1
    else: l_wins += 1
    bg   = ALT_BLUE if i % 2 == 0 else WHITE

    dat(ws2, r, 2,  m,          bold=True, bg=bg)
    dat(ws2, r, 3,  act,        fmt=EUR,   bg=bg)
    dat(ws2, r, 4,  p,          fmt=EUR,   bg=bg,
        color=GREEN if ep < 0.20 else (RED if ep > 0.40 else '000000'))
    dat(ws2, r, 5,  l,          fmt=EUR,   bg=bg,
        color=GREEN if el < 0.20 else (RED if el > 0.40 else '000000'))
    dat(ws2, r, 6,  ep,         fmt='0%',  bg=bg,
        color=GREEN if ep < 0.20 else (RED if ep > 0.40 else '000000'))
    dat(ws2, r, 7,  el,         fmt='0%',  bg=bg,
        color=GREEN if el < 0.20 else (RED if el > 0.40 else '000000'))
    b_bg  = GREEN_BG if better == 'Prophet' else LIGHT_BLUE
    b_col = GREEN    if better == 'Prophet' else MID_BLUE
    dat(ws2, r, 8,  better,     bold=True, bg=b_bg, color=b_col)
    dat(ws2, r, 9,  abs(p - l), fmt=EUR,   bg=bg)
    dat(ws2, r, 10, abs(p - l) / act, fmt='0%', bg=bg)
    ws2.row_dimensions[r].height = 18

# Totals
r_tot = N_TEST + 5
ws2.row_dimensions[r_tot].height = 22
avg_act = float(test_df['y'].mean())
avg_p   = float(prophet_test_pred.mean())
avg_l   = float(lgbm_pred_df['yhat'].mean())
avg_ep  = np.mean([abs(float(test_df['y'].iloc[i]) - float(prophet_test_pred[i])) / float(test_df['y'].iloc[i]) for i in range(N_TEST)])
avg_el  = np.mean([abs(float(test_df['y'].iloc[i]) - float(lgbm_pred_df['yhat'].iloc[i])) / float(test_df['y'].iloc[i]) for i in range(N_TEST)])

totals_row = [
    (2, 'AVERAGE',                                    None),
    (3, avg_act,                                      EUR),
    (4, avg_p,                                        EUR),
    (5, avg_l,                                        EUR),
    (6, avg_ep,                                       '0%'),
    (7, avg_el,                                       '0%'),
    (8, 'Prophet' if avg_ep < avg_el else 'LightGBM', None),
    (9, '',                                           None),
    (10, '',                                          None),
]
for col, val, fmt in totals_row:
    hdr(ws2, r_tot, col, val, bg=MID_BLUE)
    if fmt:
        ws2.cell(r_tot, col).number_format = fmt

r_leg = r_tot + 2
ws2.merge_cells(f'B{r_leg}:F{r_leg}')
ws2[f'B{r_leg}'] = 'Color guide:  Green = error < 20%   |   Red = error > 40%'
ws2[f'B{r_leg}'].font      = Font(name='Arial', size=9, italic=True, color=GRAY)
ws2[f'B{r_leg}'].alignment = Alignment(horizontal='left')

for col, w in {'B': 12, 'C': 13, 'D': 13, 'E': 14, 'F': 12, 'G': 13, 'H': 14, 'I': 13, 'J': 10}.items():
    ws2.column_dimensions[col].width = w
ws2.freeze_panes = 'B5'

# ── Sheet 3: Win Count ────────────────────────────────────────────────────────
ws3 = wb2.create_sheet('Win Count')
ws3.sheet_view.showGridLines = False

ws3.merge_cells('B2:F2')
ws3['B2'] = 'Month-by-Month Win Count'
ws3['B2'].font      = Font(name='Arial', bold=True, size=14, color=DARK_BLUE)
ws3['B2'].alignment = Alignment(horizontal='left', vertical='center')
ws3.row_dimensions[2].height = 26
ws3.row_dimensions[3].height = 10

for col, h in enumerate(['Month', 'Actual', 'Err Prophet', 'Err LightGBM', 'Winner'], 2):
    hdr(ws3, 4, col, h)
ws3.row_dimensions[4].height = 22

for i in range(N_TEST):
    r    = i + 5
    m    = test_df['ds'].iloc[i].strftime('%b %Y')
    act  = float(test_df['y'].iloc[i])
    p    = float(prophet_test_pred[i])
    l    = float(lgbm_pred_df['yhat'].iloc[i])
    ep   = abs(act - p) / act
    el   = abs(act - l) / act
    better = 'Prophet' if ep <= el else 'LightGBM'
    bg   = ALT_BLUE if i % 2 == 0 else WHITE
    b_bg  = GREEN_BG if better == 'Prophet' else LIGHT_BLUE
    b_col = GREEN    if better == 'Prophet' else MID_BLUE

    dat(ws3, r, 2, m,      bold=True, bg=bg)
    dat(ws3, r, 3, act,    fmt=EUR,   bg=bg)
    dat(ws3, r, 4, ep,     fmt='0%',  bg=bg,
        color=GREEN if ep < 0.20 else (RED if ep > 0.40 else '000000'))
    dat(ws3, r, 5, el,     fmt='0%',  bg=bg,
        color=GREEN if el < 0.20 else (RED if el > 0.40 else '000000'))
    dat(ws3, r, 6, better, bold=True, bg=b_bg, color=b_col)
    ws3.row_dimensions[r].height = 18

r_s = N_TEST + 6
ws3.merge_cells(f'B{r_s}:F{r_s}')
ws3[f'B{r_s}'] = 'Summary'
ws3[f'B{r_s}'].font      = Font(name='Arial', bold=True, size=11, color=DARK_BLUE)
ws3[f'B{r_s}'].fill      = PatternFill('solid', start_color=LIGHT_BLUE)
ws3[f'B{r_s}'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws3[f'B{r_s}'].border    = make_border()
ws3.row_dimensions[r_s].height = 20

summary_data = [
    ('Prophet wins',  p_wins,               GREEN_BG,    GREEN),
    ('LightGBM wins', l_wins,               LIGHT_BLUE,  MID_BLUE),
    ('Total months',  N_TEST,               ALT_BLUE,    DARK_BLUE),
    ('Prophet rate',  f'{p_wins/N_TEST:.0%}',GREEN_BG,   GREEN),
    ('LightGBM rate', f'{l_wins/N_TEST:.0%}',LIGHT_BLUE, MID_BLUE),
]
for col, (label, val, bg, col_c) in enumerate(summary_data, 2):
    hdr(ws3, r_s+1, col, label, bg=MID_BLUE, size=10)
    dat(ws3, r_s+2, col, val,   bold=True, bg=bg, color=col_c)
    ws3.row_dimensions[r_s+1].height = 20
    ws3.row_dimensions[r_s+2].height = 22

for col, w in {'B': 12, 'C': 13, 'D': 13, 'E': 14, 'F': 13}.items():
    ws3.column_dimensions[col].width = w
ws3.freeze_panes = 'B5'

wb2.save('LawFirm_Forecast_Results.xlsx')
print("Saved: LawFirm_Forecast_Results.xlsx")
print("\nDONE. Both files saved in the same folder as this script.")
